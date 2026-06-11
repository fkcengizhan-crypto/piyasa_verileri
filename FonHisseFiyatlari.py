import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import json  # JSON dosyası için eklendi

# ========== HİSSE FONKSİYONLARI (TradingView) ==========
def fetch_all_stocks():
    url = "https://scanner.tradingview.com/turkey/scan?label-product=screener-stock"
    headers = {"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}

    payload = {
        "columns": ["name", "description", "close"],
        "filter": [],
        "ignore_unknown_fields": False,
        "options": {"lang": "tr"},
        "range": [0, 999],
        "sort": {"sortBy": "market_cap_basic", "sortOrder": "desc"},
        "markets": ["turkey"],
        "filter2": {
            "operator": "and",
            "operands": [
                {
                    "operation": {
                        "operator": "or",
                        "operands": [
                            {
                                "operation": {
                                    "operator": "and",
                                    "operands": [
                                        {"expression": {"left": "type", "operation": "equal", "right": "stock"}},
                                        {"expression": {"left": "typespecs", "operation": "has", "right": ["common"]}}
                                    ]
                                }
                            },
                            {
                                "operation": {
                                    "operator": "and",
                                    "operands": [
                                        {"expression": {"left": "type", "operation": "equal", "right": "stock"}},
                                        {"expression": {"left": "typespecs", "operation": "has", "right": ["preferred"]}}
                                    ]
                                }
                            },
                            {
                                "operation": {
                                    "operator": "and",
                                    "operands": [
                                        {"expression": {"left": "type", "operation": "equal", "right": "dr"}}
                                    ]
                                }
                            },
                            {
                                "operation": {
                                    "operator": "and",
                                    "operands": [
                                        {"expression": {"left": "type", "operation": "equal", "right": "fund"}},
                                        {"expression": {"left": "typespecs", "operation": "has_none_of", "right": ["etf", "mutual"]}}
                                    ]
                                }
                            }
                        ]
                    }
                },
                {
                    "expression": {"left": "typespecs", "operation": "has_none_of", "right": ["pre-ipo"]}
                }
            ]
        }
    }

    all_stocks = []

    def extract_stocks(data):
        count_before = len(all_stocks)
        for item in data.get("data", []):
            try:
                d = item.get("d", [])
                if len(d) < 3:
                    continue
                if isinstance(d[0], str) and d[0]:
                    code = d[0]
                    name = d[1] if d[1] else ""
                    price = d[2]
                    all_stocks.append({
                        "Hisse Kodu": code,
                        "Hisse Adı": name,
                        "Hisse Fiyatı": price
                    })
            except Exception as e:
                print(f"  ⚠ Hisse atlandı: {e}")
        added = len(all_stocks) - count_before
        if added > 0:
            print(f"  → {added} adet hisse bulundu.")

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        total = data.get("totalCount", 0)
        extract_stocks(data)

        page_size = 999
        for start in range(page_size, total, page_size):
            payload["range"] = [start, start + page_size - 1]
            print(f"Sayfa {start // page_size + 1} çekiliyor...", end=" ")
            try:
                resp = requests.post(url, json=payload, headers=headers, timeout=30)
                resp.raise_for_status()
                page_data = resp.json()
                extract_stocks(page_data)
            except Exception as e:
                print(f"\n  ⚠ Sayfa hatası: {e}")

    except Exception as e:
        print(f"Ana hata: {e}")
        if all_stocks:
            print(f"{len(all_stocks)} hisseyle devam ediliyor...")
            return all_stocks
        return None

    return all_stocks if all_stocks else None

# ========== FON FONKSİYONLARI (TEFAS) ==========
def fon_turu_esleme_al():
    url = "https://www.tefas.gov.tr/api/fund-returns/export"
    payload = {
        "format": "json",
        "listingType": "return",
        "fundType": "YAT",
        "locale": "tr",
        "filters": {
            "kurucuKodu": None, "fonTurKod": None, "fonGrubu": None,
            "fonTurAciklama": None, "sfonTurKod": None, "islem": 1,
            "basTarih": "20260601", "bitTarih": "20260601",
            "calismaTipi": 1,
            "donemGetiri1a": "0", "donemGetiri3a": "0", "donemGetiri6a": "0",
            "donemGetiriyb": "0", "donemGetiri1y": "0", "donemGetiri3y": "0",
            "donemGetiri5y": "0", "getiriOrani": "1"
        },
        "columns": ["fonKodu", "fonTurAciklama"]
    }
    headers = {"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        df = pd.DataFrame(data)
        df = df[["fonKodu", "fonTurAciklama"]].drop_duplicates()
        return dict(zip(df["fonKodu"], df["fonTurAciklama"]))
    except Exception as e:
        print(f"Fon türü alınamadı: {e}")
        return {}

def son_fiyatlari_cek(tarih_str):
    url = "https://www.tefas.gov.tr/api/funds/fonGnlBlgSiraliGetir"
    headers = {"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}
    payload = {
        "fonTipi": "YAT",
        "basTarih": tarih_str,
        "bitTarih": tarih_str,
        "basSira": 1,
        "bitSira": 5000,
        "dil": "TR"
    }
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        return data.get("resultList", [])
    except Exception as e:
        print(f"Fiyat çekme hatası: {e}")
        return []

def fetch_all_funds():
    tarih_denenecek = []
    bugun = datetime.now().date()
    for i in range(3):
        tarih_denenecek.append(bugun.strftime("%Y%m%d"))
        bugun = bugun - timedelta(days=1)

    tur_map = fon_turu_esleme_al()

    sonuclar = []
    for tarih_str in tarih_denenecek:
        veri = son_fiyatlari_cek(tarih_str)
        if veri:
            print(f"  → {len(veri)} fon bulundu.")
            sonuclar = veri
            break
        else:
            print("✗ Veri yok.")

    if not sonuclar:
        print("Son 3 günde hiç veri bulunamadı.")
        return None

    df = pd.DataFrame(sonuclar)
    df = df[["fonKodu", "fonUnvan", "fiyat"]].rename(columns={
        "fonKodu": "Fon Kodu",
        "fonUnvan": "Fon Adı",
        "fiyat": "Fon Fiyatı"
    })
    df["Fon Türü"] = df["Fon Kodu"].map(tur_map).fillna("Bilinmiyor")

    df = df[df["Fon Türü"] != "Bilinmiyor"]
    df = df[~df["Fon Türü"].str.contains("Serbest", na=False)]

    if df.empty:
        print("Filtreleme sonrası hiç fon kalmadı.")
        return None

    df = df[["Fon Kodu", "Fon Adı", "Fon Türü", "Fon Fiyatı"]]
    return df

# ========== ANA PROGRAM ==========
if __name__ == "__main__":
    print("=== BIST HİSSELERİ ÇEKİLİYOR ===")
    stocks = fetch_all_stocks()
    print("\n=== FON FİYATLARI ÇEKİLİYOR ===")
    funds = fetch_all_funds()

    if stocks is None and funds is None:
        print("Hiç veri alınamadı, çıkılıyor.")
        exit()

    dosya = "FonHisseFiyatlari.xlsx"
    
    # Excel yazma
    with pd.ExcelWriter(dosya, engine='openpyxl') as writer:
        if funds is not None:
            funds.to_excel(writer, sheet_name="Fiyatlar", startrow=0, startcol=0, index=False)
            fund_cols = 4
        else:
            fund_cols = 0
        
        if stocks is not None:
            startcol_stocks = fund_cols + 1 if fund_cols > 0 else 0
            df_stocks = pd.DataFrame(stocks)
            df_stocks = df_stocks[["Hisse Kodu", "Hisse Adı", "Hisse Fiyatı"]]
            df_stocks.to_excel(writer, sheet_name="Fiyatlar", startrow=0, startcol=startcol_stocks, index=False)
            stocks_startcol = startcol_stocks
        else:
            stocks_startcol = None

    # Tablo stilleri ekleme
    wb = load_workbook(dosya)
    ws = wb["Fiyatlar"]

    if funds is not None:
        max_row_funds = len(funds) + 1
        tablo_ref_funds = f"A1:{get_column_letter(4)}{max_row_funds}"
        tbl_funds = Table(displayName="tbl_fonfiyatlari", ref=tablo_ref_funds)
        tbl_funds.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(tbl_funds)
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12

    if stocks is not None and stocks_startcol is not None:
        first_col_letter = get_column_letter(stocks_startcol + 1)
        last_col_letter = get_column_letter(stocks_startcol + 3)
        max_row_stocks = len(stocks) + 1
        tablo_ref_stocks = f"{first_col_letter}1:{last_col_letter}{max_row_stocks}"
        tbl_stocks = Table(displayName="tbl_hissefiyatlari", ref=tablo_ref_stocks)
        tbl_stocks.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(tbl_stocks)
        ws.column_dimensions[first_col_letter].width = 12
        ws.column_dimensions[get_column_letter(stocks_startcol + 2)].width = 40
        ws.column_dimensions[last_col_letter].width = 14

    wb.save(dosya)
    print(f"\n✔ '{dosya}' kaydedildi.")

    # ========== JSON DOSYASI OLUŞTURMA ==========
    json_data = {
        "guncelleme_zamani": datetime.now().isoformat(),
        "fonlar": funds.to_dict(orient='records') if funds is not None else [],
        "hisseler": stocks if stocks is not None else []
    }

    json_dosya = "FonHisseFiyatlari.json"
    with open(json_dosya, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)

    print(f"✔ '{json_dosya}' kaydedildi.")