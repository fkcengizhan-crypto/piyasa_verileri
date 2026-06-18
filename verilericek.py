import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import json
import re
from bs4 import BeautifulSoup


# ========== HİSSE FONKSİYONLARI (TradingView) ==========
def fetch_all_stocks():
    url = "https://scanner.tradingview.com/turkey/scan?label-product=screener-stock"
    headers = {"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}

    payload = {
        "columns": ["name", "description", "close"],
        "filter": [],
        "ignore_unknown_fields": False,
        "options": {"lang": "tr"},
        "range": [0, 999],
        "sort": {"sortBy": "market_cap_basic", "sortOrder": "desc"},
        "markets": ["turkey"],
        "filter2": {
            "operator": "and",
            "operands": [
                {
                    "operation": {
                        "operator": "or",
                        "operands": [
                            {
                                "operation": {
                                    "operator": "and",
                                    "operands": [
                                        {"expression": {"left": "type", "operation": "equal", "right": "stock"}},
                                        {"expression": {"left": "typespecs", "operation": "has", "right": ["common"]}}
                                    ]
                                }
                            },
                            {
                                "operation": {
                                    "operator": "and",
                                    "operands": [
                                        {"expression": {"left": "type", "operation": "equal", "right": "stock"}},
                                        {"expression": {"left": "typespecs", "operation": "has", "right": ["preferred"]}}
                                    ]
                                }
                            },
                            {
                                "operation": {
                                    "operator": "and",
                                    "operands": [
                                        {"expression": {"left": "type", "operation": "equal", "right": "dr"}}
                                    ]
                                }
                            },
                            {
                                "operation": {
                                    "operator": "and",
                                    "operands": [
                                        {"expression": {"left": "type", "operation": "equal", "right": "fund"}},
                                        {"expression": {"left": "typespecs", "operation": "has_none_of", "right": ["etf", "mutual"]}}
                                    ]
                                }
                            }
                        ]
                    }
                },
                {
                    "expression": {"left": "typespecs", "operation": "has_none_of", "right": ["pre-ipo"]}
                }
            ]
        }
    }

    all_stocks = []

    def extract_stocks(data):
        count_before = len(all_stocks)
        for item in data.get("data", []):
            try:
                d = item.get("d", [])
                if len(d) < 3:
                    continue
                if isinstance(d[0], str) and d[0]:
                    code = d[0]
                    name = d[1] if d[1] else ""
                    price = d[2]
                    all_stocks.append({
                        "Hisse Kodu": code,
                        "Hisse Adı": name,
                        "Hisse Fiyatı": price
                    })
            except Exception as e:
                print(f"  ⚠ Hisse atlandı: {e}")
        added = len(all_stocks) - count_before
        if added > 0:
            print(f"  → {added} adet hisse bulundu.")

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        total = data.get("totalCount", 0)
        extract_stocks(data)

        page_size = 999
        for start in range(page_size, total, page_size):
            payload["range"] = [start, start + page_size - 1]
            print(f"Sayfa {start // page_size + 1} çekiliyor...", end=" ")
            try:
                resp = requests.post(url, json=payload, headers=headers, timeout=30)
                resp.raise_for_status()
                page_data = resp.json()
                extract_stocks(page_data)
            except Exception as e:
                print(f"\n  ⚠ Sayfa hatası: {e}")

    except Exception as e:
        print(f"Ana hata: {e}")
        if all_stocks:
            print(f"{len(all_stocks)} hisseyle devam ediliyor...")
            return all_stocks
        return None

    return all_stocks if all_stocks else None


# ========== FON FONKSİYONLARI (TEFAS) ==========
def fon_turu_esleme_al():
    url = "https://www.tefas.gov.tr/api/fund-returns/export"
    payload = {
        "format": "json",
        "listingType": "return",
        "fundType": "YAT",
        "locale": "tr",
        "filters": {
            "kurucuKodu": None, "fonTurKod": None, "fonGrubu": None,
            "fonTurAciklama": None, "sfonTurKod": None, "islem": 1,
            "basTarih": "20260601", "bitTarih": "20260601",
            "calismaTipi": 1,
            "donemGetiri1a": "0", "donemGetiri3a": "0", "donemGetiri6a": "0",
            "donemGetiriyb": "0", "donemGetiri1y": "0", "donemGetiri3y": "0",
            "donemGetiri5y": "0", "getiriOrani": "1"
        },
        "columns": ["fonKodu", "fonTurAciklama"]
    }
    headers = {"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        df = pd.DataFrame(data)
        df = df[["fonKodu", "fonTurAciklama"]].drop_duplicates()
        return dict(zip(df["fonKodu"], df["fonTurAciklama"]))
    except Exception as e:
        print(f"Fon türü alınamadı: {e}")
        return {}

def son_fiyatlari_cek(tarih_str):
    url = "https://www.tefas.gov.tr/api/funds/fonGnlBlgSiraliGetir"
    headers = {"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}
    payload = {
        "fonTipi": "YAT",
        "basTarih": tarih_str,
        "bitTarih": tarih_str,
        "basSira": 1,
        "bitSira": 5000,
        "dil": "TR"
    }
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        return data.get("resultList", [])
    except Exception as e:
        print(f"Fiyat çekme hatası: {e}")
        return []

def fetch_all_funds():
    tarih_denenecek = []
    bugun = datetime.now().date()
    for i in range(3):
        tarih_denenecek.append(bugun.strftime("%Y%m%d"))
        bugun = bugun - timedelta(days=1)

    tur_map = fon_turu_esleme_al()

    sonuclar = []
    for tarih_str in tarih_denenecek:
        veri = son_fiyatlari_cek(tarih_str)
        if veri:
            print(f"  → {len(veri)} fon bulundu.")
            sonuclar = veri
            break
        else:
            print("✗ Veri yok.")

    if not sonuclar:
        print("Son 3 günde hiç veri bulunamadı.")
        return None

    df = pd.DataFrame(sonuclar)
    df = df[["fonKodu", "fonUnvan", "fiyat"]].rename(columns={
        "fonKodu": "Fon Kodu",
        "fonUnvan": "Fon Adı",
        "fiyat": "Fon Fiyatı"
    })
    df["Fon Türü"] = df["Fon Kodu"].map(tur_map).fillna("Bilinmiyor")

    df = df[df["Fon Türü"] != "Bilinmiyor"]
    df = df[~df["Fon Türü"].str.contains("Serbest", na=False)]

    if df.empty:
        print("Filtreleme sonrası hiç fon kalmadı.")
        return None

    df = df[["Fon Kodu", "Fon Adı", "Fon Türü", "Fon Fiyatı"]]
    return df


# ========== DOVIZ.COM VERİLERİ (STATİK HTML) ==========
def format_price(price_str):
    """
    Ham fiyat string'ini düzenler:
    - Eğer zaten virgül ile ondalık ayrılmışsa (örn: 14.421,15) aynen döndür.
    - Eğer nokta ile ayrılmışsa son noktayı virgüle çevir (örn: 14.421.15 -> 14.421,15)
    - Eğer sadece nokta yoksa aynen döndür (örn: 65.820 -> 65.820)
    """
    if not price_str:
        return "Veri yok"
    
    # Önce varsa dolar işaretini temizle
    price_str = price_str.replace('$', '').strip()
    
    # Eğer zaten virgül varsa (Türk formatı) direkt döndür
    if ',' in price_str:
        return price_str
    
    # Nokta ile ayrılmışsa son noktayı virgül yap
    parts = price_str.split('.')
    if len(parts) > 1:
        # Son parça ondalık kısım, diğerleri binlik
        integer_part = '.'.join(parts[:-1])
        decimal_part = parts[-1]
        return f"{integer_part},{decimal_part}"
    else:
        # Nokta yoksa aynen döndür (tam sayı)
        return price_str

def fetch_doviz_data():
    """
    Doviz.com'dan istenen tüm verileri çeker.
    Döner: List[Dict] - {'Sembol': ..., 'Fiyat': ...}
    """
    
    # URL ve sembol eşleştirmesi - İstenen sıraya göre düzenlendi
    urls = {
        "BIST 100": "https://borsa.doviz.com/endeksler/xu100-bist-100",
        "Dolar": "https://kur.doviz.com/serbest-piyasa/amerikan-dolari",
        "Euro": "https://kur.doviz.com/serbest-piyasa/euro",
        "BTCUSD": "https://www.doviz.com/kripto-paralar/bitcoin",
        "Tahvil": "https://www.doviz.com/tahvil/tr-5-yillik-tahvil",
        "Brent Petrol": "https://www.doviz.com/emtia/brent-petrol",
        "Gr Altın": "https://altin.doviz.com/gram-altin",
        "Çey.Altın": "https://altin.doviz.com/ceyrek-altin",
        "Tam Altın": "https://altin.doviz.com/tam-altin",
        "14 Ayar Bilezik": "https://altin.doviz.com/14-ayar-altin",
        "18 Ayar Bilezik": "https://altin.doviz.com/18-ayar-altin",
        "22 Ayar Bilezik": "https://altin.doviz.com/22-ayar-bilezik",
        "Ons Altın": "https://altin.doviz.com/ons"
    }
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    sonuclar = []
    
    for sembol, url in urls.items():
        try:
            print(f"  → {sembol} çekiliyor...", end=" ")
            resp = requests.get(url, headers=headers, timeout=15)
            resp.raise_for_status()
            
            soup = BeautifulSoup(resp.text, 'html.parser')
            fiyat_raw = None
            
            # 1. Yöntem: data-socket-key ve data-socket-attr="s" olan div'leri ara
            div = soup.find('div', {'data-socket-key': True, 'data-socket-attr': 's'})
            if div:
                fiyat_raw = div.text.strip()
            else:
                # 2. Yöntem: Sayfa başlığından veya h1'den fiyat bilgisini çek
                h1 = soup.find('h1')
                if h1:
                    h1_text = h1.text
                    match = re.search(r'Son\s*\([^)]*\)\s*([\d.,]+)', h1_text)
                    if match:
                        fiyat_raw = match.group(1)
                
                # 3. Yöntem: Meta description'dan dene
                if not fiyat_raw:
                    meta = soup.find('meta', {'name': 'description'})
                    if meta and meta.get('content'):
                        content = meta.get('content')
                        match = re.search(r'([\d.,]+)\s*(?:seviyesinde|dolar|TL|$)', content)
                        if match:
                            fiyat_raw = match.group(1)
                
                # 4. Yöntem: Sayfa başlığı (title) içinden fiyat ara
                if not fiyat_raw:
                    title = soup.find('title')
                    if title:
                        title_text = title.text
                        match = re.search(r'([\d.,]+)\s*(?:TL|$)', title_text)
                        if match:
                            fiyat_raw = match.group(1)
            
            if fiyat_raw:
                # Formatı düzelt (dolar işareti temizlenir, nokta/virgül düzenlenir)
                fiyat = format_price(fiyat_raw)
                sonuclar.append({"Sembol": sembol, "Fiyat": fiyat})
                print("✅")
            else:
                sonuclar.append({"Sembol": sembol, "Fiyat": "Bulunamadı"})
                print("❌ (bulunamadı)")
                
        except Exception as e:
            print(f"❌ Hata: {str(e)[:50]}")
            sonuclar.append({"Sembol": sembol, "Fiyat": "Hata"})
    
    return sonuclar


# ========== ANA PROGRAM ==========
if __name__ == "__main__":
    print("=== BIST HİSSELERİ ÇEKİLİYOR ===")
    stocks = fetch_all_stocks()
    
    print("\n=== FON FİYATLARI ÇEKİLİYOR ===")
    funds = fetch_all_funds()
    
    print("\n=== DOVIZ.COM VERİLERİ ÇEKİLİYOR ===")
    doviz = fetch_doviz_data()
    if doviz:
        print(f"  → {len(doviz)} veri bulundu.")
    else:
        print("  ⚠ Doviz.com verisi alınamadı.")

    if stocks is None and funds is None and doviz is None:
        print("Hiç veri alınamadı, çıkılıyor.")
        exit()

    dosya = "piyasa_verileri.xlsx"
    
    stocks_startcol = None
    doviz_startcol = None

    # Excel yazma (Tablolar arasında 1 sütun boşluk)
    with pd.ExcelWriter(dosya, engine='openpyxl') as writer:
        current_col = 0
        
        if funds is not None:
            funds.to_excel(writer, sheet_name="Fiyatlar", startrow=0, startcol=current_col, index=False)
            current_col += 4  # Fon Kodu, Fon Adı, Fon Türü, Fon Fiyatı
            current_col += 1  # ⬅️ BOŞ SÜTUN
        
        if stocks is not None:
            stocks_startcol = current_col
            df_stocks = pd.DataFrame(stocks)
            df_stocks = df_stocks[["Hisse Kodu", "Hisse Adı", "Hisse Fiyatı"]]
            df_stocks.to_excel(writer, sheet_name="Fiyatlar", startrow=0, startcol=current_col, index=False)
            current_col += 3
            current_col += 1  # ⬅️ BOŞ SÜTUN
            
        if doviz is not None:
            doviz_startcol = current_col
            df_doviz = pd.DataFrame(doviz)
            df_doviz.to_excel(writer, sheet_name="Fiyatlar", startrow=0, startcol=current_col, index=False)

    # Tablo stilleri ve biçimlendirme
    wb = load_workbook(dosya)
    ws = wb["Fiyatlar"]

    # 1. FON TABLOSU
    if funds is not None:
        max_row_funds = len(funds) + 1
        tablo_ref_funds = f"A1:{get_column_letter(4)}{max_row_funds}"
        tbl_funds = Table(displayName="tbl_fonfiyatlari", ref=tablo_ref_funds)
        tbl_funds.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(tbl_funds)
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        # Fon fiyat sütunu (D) sağa yasla
        for row in range(2, max_row_funds + 1):
            cell = ws[f'D{row}']
            cell.alignment = Alignment(horizontal='right')

    # 2. HİSSE TABLOSU
    if stocks is not None and stocks_startcol is not None:
        first_col_letter = get_column_letter(stocks_startcol + 1)
        last_col_letter = get_column_letter(stocks_startcol + 3)
        max_row_stocks = len(stocks) + 1
        tablo_ref_stocks = f"{first_col_letter}1:{last_col_letter}{max_row_stocks}"
        tbl_stocks = Table(displayName="tbl_hissefiyatlari", ref=tablo_ref_stocks)
        tbl_stocks.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(tbl_stocks)
        ws.column_dimensions[first_col_letter].width = 12
        ws.column_dimensions[get_column_letter(stocks_startcol + 2)].width = 40
        ws.column_dimensions[last_col_letter].width = 14
        # Hisse fiyat sütunu (3. sütun) sağa yasla
        price_col_letter = get_column_letter(stocks_startcol + 3)
        for row in range(2, max_row_stocks + 1):
            cell = ws[f'{price_col_letter}{row}']
            cell.alignment = Alignment(horizontal='right')

    # 3. DOVIZ TABLOSU (sadece 2 sütun: Sembol ve Fiyat)
    if doviz is not None and doviz_startcol is not None:
        first_col_letter = get_column_letter(doviz_startcol + 1)
        last_col_letter = get_column_letter(doviz_startcol + 2)
        max_row_doviz = len(doviz) + 1
        tablo_ref_doviz = f"{first_col_letter}1:{last_col_letter}{max_row_doviz}"
        tbl_doviz = Table(displayName="tbl_doviz", ref=tablo_ref_doviz)
        tbl_doviz.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(tbl_doviz)
        ws.column_dimensions[first_col_letter].width = 18
        ws.column_dimensions[last_col_letter].width = 16
        # Fiyat sütunu (2. sütun) sağa yasla
        price_col_letter = get_column_letter(doviz_startcol + 2)
        for row in range(2, max_row_doviz + 1):
            cell = ws[f'{price_col_letter}{row}']
            cell.alignment = Alignment(horizontal='right')

    # Boş sütun genişliklerini ayarla
    for col_idx in [5, 9]:
        try:
            ws.column_dimensions[get_column_letter(col_idx)].width = 3
        except:
            pass

    wb.save(dosya)
    print(f"\n✔ '{dosya}' kaydedildi.")

    # ========== JSON DOSYASI ==========
    json_data = {
        "guncelleme_zamani": datetime.now().isoformat(),
        "fonlar": funds.to_dict(orient='records') if funds is not None else [],
        "hisseler": stocks if stocks is not None else [],
        "doviz": doviz if doviz is not None else []
    }

    json_dosya = "piyasa_verileri.json"
    with open(json_dosya, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)

    print(f"✔ '{json_dosya}' kaydedildi.")